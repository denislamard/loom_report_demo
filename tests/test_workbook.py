"""Le classeur : gabarits de formule, contrat de sélection, structure produite.

Les assertions portent sur les **chaînes de formule**, jamais sur des valeurs.
C'est ce qui permet de tester sans LibreOffice, donc en intégration continue en
quelques secondes. La concordance des valeurs est contrôlée séparément par un
recalcul complet, hors CI.

Deux familles méritent l'attention.

La **résolution des colonnes** : aucune lettre n'est codée en dur, elles sont
dérivées des en-têtes CSV. Un test vérifie qu'une formule pointe bien la colonne
du champ demandé — une `SUMIFS` décalée d'une colonne ne lève aucune erreur, elle
rend un mauvais chiffre.

La **parité pandas / Excel** : les colonnes calculées existent deux fois, une
fois en pandas pour l'analyse, une fois en formules pour le classeur. Rien
n'oblige mécaniquement les deux jeux à porter les mêmes noms. Un test les
confronte.
"""

from __future__ import annotations

import json
from functools import cache
from pathlib import Path

import pytest
from openpyxl import load_workbook

from loom_report_demo import paths
from loom_report_demo.analysis import catalogue as cat
from loom_report_demo.analysis.chargement import COLONNE_DATE, Donnees, charger
from loom_report_demo.niveaux import NIVEAUX, Niveau
from loom_report_demo.workbook import ORDRE, construire
from loom_report_demo.workbook import formules as f
from loom_report_demo.workbook.feuilles_donnees import ecrire_feuilles
from loom_report_demo.workbook.schema import FEUILLE_DE_BASE, construire_schemas
from loom_report_demo.parsing import charger as charger_selection
from loom_report_demo.workbook.selection import Indicateur, Selection

FIXTURE = Path(__file__).parent / "fixtures" / "gestion.json"


@cache
def schemas():  # noqa: ANN201
    return construire_schemas(f.COLONNES_CALCULEES)


@cache
def jeu() -> Donnees:
    return charger()


@cache
def selection() -> Selection:
    return charger_selection(FIXTURE)


@cache
def classeur_construit(dossier: str):  # noqa: ANN201
    chemin = Path(dossier) / "rapport.xlsx"
    construire(selection(), chemin, source=jeu())
    return load_workbook(chemin)


# ------------------------------------------------- résolution des colonnes
def test_les_lettres_viennent_des_entetes_csv() -> None:
    devis = schemas()["Devis"]
    assert devis.lettre("devis_id") == "A"
    assert devis.plage("montant_ht").startswith("Devis!$")


def test_un_champ_inconnu_liste_les_disponibles() -> None:
    with pytest.raises(KeyError, match="montant_ht"):
        schemas()["Devis"].lettre("chiffre")


def test_les_colonnes_calculees_prolongent_les_donnees_sources() -> None:
    devis = schemas()["Devis"]
    calculees = [c.cle for c in f.COLONNES_CALCULEES["Devis"]]
    assert devis.champs[-len(calculees) :] == tuple(calculees)


@pytest.mark.parametrize("feuille", ["Devis", "Factures", "Interventions"])
def test_parite_des_colonnes_calculees_avec_pandas(feuille: str) -> None:
    """Les mêmes colonnes existent des deux côtés, sous les mêmes noms."""
    base = next(b for b, nom in FEUILLE_DE_BASE.items() if nom == feuille)
    colonnes_pandas = set(jeu().table(base).columns)
    for colonne in f.COLONNES_CALCULEES[feuille]:
        assert colonne.cle in colonnes_pandas, f"{feuille}.{colonne.cle} absente du chargement"


def test_toute_colonne_utilisee_par_le_catalogue_existe_dans_le_classeur() -> None:
    for mesure in cat.MESURES.values():
        if mesure.agregat is None or mesure.agregat.distinct is not None:
            continue
        schema = schemas()[FEUILLE_DE_BASE[mesure.base]]
        schema.lettre(mesure.agregat.numerateur)
        if mesure.agregat.denominateur:
            schema.lettre(mesure.agregat.denominateur)


# ---------------------------------------------------- gabarits de formule
def test_une_somme_produit_un_sumifs_borne_par_la_fenetre() -> None:
    mesure = cat.mesure("ca_facture_ht")
    schema = schemas()["Factures"]
    formule = f.formule_mesure(mesure, schema, COLONNE_DATE[cat.Base.FACTURES])
    assert formule.startswith("=SUMIFS(")
    assert schema.plage("montant_ht") in formule
    assert f'">="&{f.DEBUT}' in formule
    assert f'"<="&{f.FIN}' in formule


def test_un_ratio_divise_deux_sommes_et_jamais_des_moyennes() -> None:
    """La moyenne de trois taux de marge n'est pas le taux de marge de trois agences."""
    mesure = cat.mesure("taux_marge_brute")
    schema = schemas()["Factures"]
    formule = f.formule_mesure(mesure, schema, COLONNE_DATE[cat.Base.FACTURES])
    assert formule.count("SUMIFS(") == 2
    assert "AVERAGE" not in formule
    assert schema.plage("marge") in formule
    assert schema.plage("montant_ht") in formule


def test_le_facteur_du_dso_apparait_dans_la_formule() -> None:
    formule = f.formule_mesure(
        cat.mesure("dso"), schemas()["Factures"], COLONNE_DATE[cat.Base.FACTURES]
    )
    assert "365.0*(" in formule


def test_le_decalage_de_la_derive_apparait_dans_la_formule() -> None:
    formule = f.formule_mesure(
        cat.mesure("taux_derive_horaire"),
        schemas()["Interventions"],
        COLONNE_DATE[cat.Base.INTERVENTIONS],
    )
    assert formule.rstrip(')"').endswith("-1.0") or "-1.0," in formule


def test_une_ventilation_ajoute_un_critere_de_dimension() -> None:
    schema = schemas()["Devis"]
    formule = f.formule_mesure(
        cat.mesure("taux_transformation"),
        schema,
        COLONNE_DATE[cat.Base.DEVIS],
        dimension=("tranche_relance", '"0-3 j"'),
    )
    assert schema.plage("tranche_relance") in formule
    assert '"0-3 j"' in formule


def test_la_fenetre_de_comparaison_pointe_dautres_cellules() -> None:
    schema = schemas()["Factures"]
    actuelle = f.formule_mesure(
        cat.mesure("ca_facture_ht"), schema, COLONNE_DATE[cat.Base.FACTURES]
    )
    passee = f.formule_mesure(
        cat.mesure("ca_facture_ht"),
        schema,
        COLONNE_DATE[cat.Base.FACTURES],
        f.DEBUT_COMPARAISON,
        f.FIN_COMPARAISON,
    )
    assert actuelle != passee
    assert f.DEBUT_COMPARAISON in passee


def test_une_mesure_speciale_annonce_son_absence_de_gabarit() -> None:
    with pytest.raises(NotImplementedError, match="jalon 7"):
        f.formule_mesure(
            cat.mesure("concentration_client"),
            schemas()["Factures"],
            COLONNE_DATE[cat.Base.FACTURES],
        )


def test_un_denominateur_distinct_annonce_son_absence_de_gabarit() -> None:
    with pytest.raises(NotImplementedError, match="jalon 7"):
        f.formule_mesure(
            cat.mesure("ca_par_technicien"),
            schemas()["Interventions"],
            COLONNE_DATE[cat.Base.INTERVENTIONS],
        )


def test_les_seuils_de_tranches_viennent_du_chargement() -> None:
    """Une seule source de vérité : sinon pandas et Excel divergent au premier réglage."""
    from loom_report_demo.analysis.chargement import SEUIL_EXCEPTION_JOURS, TRANCHES_MONTANT

    gabarits = {c.cle: c.gabarit for c in f.COLONNES_CALCULEES["Factures"]}
    for borne, libelle in TRANCHES_MONTANT:
        assert f"<={borne}" in gabarits["tranche_montant"]
        assert libelle in gabarits["tranche_montant"]
    assert f.SEUIL_EXCEPTION in gabarits["exception"]
    assert SEUIL_EXCEPTION_JOURS == 90


# ------------------------------------------------------ contrat de sélection
def test_la_fixture_est_valide() -> None:
    selection().valider()


def test_la_selection_impose_le_nombre_dindicateurs() -> None:
    attendus = NIVEAUX[Niveau.GESTION].nb_variables
    with pytest.raises(ValueError, match=str(attendus)):
        Selection(
            niveau=Niveau.GESTION, variables=(Indicateur(mesure="ca_facture_ht"),)
        ).valider()


def test_la_selection_refuse_un_doublon() -> None:
    indicateur = Indicateur(mesure="taux_marge_brute", dimension="agence")
    with pytest.raises(ValueError, match="double"):
        Selection(niveau=Niveau.GESTION, variables=(indicateur,) * 4).valider()


def test_la_selection_refuse_une_mesure_hors_niveau() -> None:
    with pytest.raises(ValueError, match="pas éligible"):
        Selection(
            niveau=Niveau.STRATEGIQUE,
            variables=(
                Indicateur(mesure="ca_devise_ht", dimension="agence"),
                Indicateur(mesure="taux_marge_brute", dimension="agence"),
            ),
        ).valider()


def test_la_selection_refuse_un_croisement_tautologique() -> None:
    with pytest.raises(ValueError, match="tautologique"):
        Selection(
            niveau=Niveau.GESTION,
            variables=(
                Indicateur(mesure="panier_moyen_gagne", dimension="tranche_montant"),
                Indicateur(mesure="taux_marge_brute", dimension="agence"),
                Indicateur(mesure="taux_transformation", dimension="agence"),
                Indicateur(mesure="taux_derive_horaire", dimension="agence"),
            ),
        ).valider()


def test_la_selection_refuse_deux_hypotheses_de_meme_identifiant() -> None:
    """Le décodage relève du parsing depuis le jalon 5 ; la garde reste la même."""
    from loom_report_demo.parsing import ErreurSortie, analyser
    from loom_report_demo.workbook.selection import HypothesePerdue

    contenu = json.loads(FIXTURE.read_text(encoding="utf-8"))
    contenu["hypotheses"][1]["identifiant"] = contenu["hypotheses"][0]["identifiant"]
    with pytest.raises(ErreurSortie, match="identifiant"):
        analyser(contenu)
    assert HypothesePerdue("H1", "a", "b").identifiant == "H1"


def test_le_socle_nest_jamais_choisi_par_lagent() -> None:
    """Il vient du niveau, pas de la sélection : c'est ce qui rend les éditions comparables."""
    assert selection().socle == NIVEAUX[Niveau.GESTION].socle


# --------------------------------------------------------- classeur produit
def test_toutes_les_feuilles_attendues_sont_presentes(tmp_path: Path) -> None:
    assert classeur_construit(str(tmp_path)).sheetnames == list(ORDRE)


def test_la_restitution_precede_les_sources(tmp_path: Path) -> None:
    noms = classeur_construit(str(tmp_path)).sheetnames
    assert noms.index("Synthèse") < noms.index("Devis")


def test_aucune_formule_ne_porte_de_reference_cassee(tmp_path: Path) -> None:
    classeur = classeur_construit(str(tmp_path))
    for feuille in classeur.worksheets:
        for rangee in feuille.iter_rows():
            for cellule in rangee:
                if isinstance(cellule.value, str) and cellule.value.startswith("="):
                    assert "#REF!" not in cellule.value, f"{feuille.title}!{cellule.coordinate}"


def test_toute_feuille_referencee_existe(tmp_path: Path) -> None:
    import re

    classeur = classeur_construit(str(tmp_path))
    connues = set(classeur.sheetnames)
    motif = re.compile(r"([A-Za-zÀ-ÿ][A-Za-zÀ-ÿ' ]*)!\$")
    for feuille in classeur.worksheets:
        for rangee in feuille.iter_rows():
            for cellule in rangee:
                if isinstance(cellule.value, str) and cellule.value.startswith("="):
                    for nom in motif.findall(cellule.value):
                        assert nom in connues, f"{feuille.title} référence {nom!r}"


def test_le_classeur_est_massivement_formule(tmp_path: Path) -> None:
    """Aucune valeur figée : c'est la promesse du rapport."""
    classeur = classeur_construit(str(tmp_path))
    total = sum(
        1
        for feuille in classeur.worksheets
        for rangee in feuille.iter_rows()
        for cellule in rangee
        if isinstance(cellule.value, str) and cellule.value.startswith("=")
    )
    assert total > 50_000, total


def test_la_date_de_situation_est_une_saisie(tmp_path: Path) -> None:
    parametres = classeur_construit(str(tmp_path))["Paramètres"]
    cellule = parametres.cell(row=f.LIGNE_SITUATION, column=3)
    assert cellule.font.color is not None
    assert cellule.font.color.rgb.endswith("0000FF")


def test_les_bornes_de_fenetre_derivent_de_la_situation(tmp_path: Path) -> None:
    """Modifier la date de situation doit recalculer tout le rapport."""
    parametres = classeur_construit(str(tmp_path))["Paramètres"]
    for decalage in range(1, 5):
        formule = parametres.cell(row=f.LIGNE_SITUATION + decalage, column=3).value
        assert isinstance(formule, str) and formule.startswith("=$C$8")


def test_chaque_indicateur_choisi_a_son_bloc(tmp_path: Path) -> None:
    feuille = classeur_construit(str(tmp_path))["Indicateurs"]
    titres = {c.value for rangee in feuille.iter_rows(min_col=2, max_col=2) for c in rangee}
    for indicateur in selection().variables:
        mesure = cat.mesure(indicateur.mesure)
        assert any(isinstance(t, str) and t.startswith(mesure.libelle) for t in titres)


def test_les_hypotheses_ecartees_sont_livrees(tmp_path: Path) -> None:
    """Elles valent autant que les trouvailles."""
    feuille = classeur_construit(str(tmp_path))["Ce qui a été regardé"]
    textes = {
        c.value for rangee in feuille.iter_rows(min_col=2, max_col=3) for c in rangee
    }
    for hypothese in selection().ecartees:
        assert hypothese.enonce in textes


def test_lempreinte_du_jeu_figure_dans_le_classeur(tmp_path: Path) -> None:
    feuille = classeur_construit(str(tmp_path))["Paramètres"]
    noms = {c.value for rangee in feuille.iter_rows(min_col=2, max_col=2) for c in rangee}
    for fichier in paths.FICHIERS_DONNEES:
        assert fichier in noms


def test_le_detail_mensuel_couvre_toute_la_periode(tmp_path: Path) -> None:
    feuille = classeur_construit(str(tmp_path))["Détail mensuel"]
    mois = [
        c.value
        for rangee in feuille.iter_rows(min_row=9, min_col=2, max_col=2)
        for c in rangee
        if isinstance(c.value, str) and c.value[:4].isdigit()
    ]
    assert len(mois) >= 46
    assert mois == sorted(mois)


def test_les_feuilles_de_donnees_portent_leurs_colonnes_calculees(tmp_path: Path) -> None:
    classeur = classeur_construit(str(tmp_path))
    for feuille, colonnes in f.COLONNES_CALCULEES.items():
        if not colonnes:
            continue
        entetes = [c.value for c in classeur[feuille][1]]
        for colonne in colonnes:
            assert colonne.entete in entetes, f"{feuille} : {colonne.entete} absente"


def test_ecrire_feuilles_ne_depend_pas_du_repertoire_courant(tmp_path: Path) -> None:
    from openpyxl import Workbook

    classeur = Workbook()
    classeur.remove(classeur.active)
    ecrire_feuilles(classeur, schemas())
    assert set(classeur.sheetnames) == set(f.COLONNES_CALCULEES)
