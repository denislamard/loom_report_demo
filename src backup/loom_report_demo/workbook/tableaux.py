"""Les feuilles de restitution : paramètres, détail mensuel, synthèse, indicateurs.

Toutes les valeurs affichées sont des formules. Les seules données figées sont
celles qui décrivent le processus lui-même — l'empreinte du jeu, la liste des
candidats examinés — parce qu'elles constatent un état passé et n'ont pas à
bouger si l'on corrige une facture.

La synthèse porte la décomposition d'écart, qui est ce qui sépare un instrument
de mesure d'un instrument de décision. Savoir que le chiffre d'affaires a
progressé de 3,5 % ne dit rien ; savoir que c'est du volume malgré un effet prix
défavorable dit quoi faire.
"""

from __future__ import annotations

from datetime import date
from math import ceil

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from loom_report_demo.analysis import catalogue as cat
from loom_report_demo.analysis.cadrages import Fenetre
from loom_report_demo.analysis.chargement import COLONNE_DATE
from loom_report_demo.analysis.criblage import Criblage
from loom_report_demo.analysis.files import File
from loom_report_demo.fingerprint import EmpreinteJeu, grouper
from loom_report_demo.workbook import formules as f
from loom_report_demo.workbook import theme
from loom_report_demo.workbook.schema import FEUILLE_DE_BASE, Schema
from loom_report_demo.workbook.selection import Indicateur, Selection

#: Hauteur d'une ligne de tableur, en centimètres. Sert à convertir la hauteur
#: d'un graphique en nombre de lignes occupées, pour que deux blocs successifs
#: ne se recouvrent jamais.
HAUTEUR_LIGNE_CM = 0.53
HAUTEUR_BARRE_CM = 1.0
HAUTEUR_SOCLE_CM = 2.5

LIGNE_MOIS = 9


def _schema_de(mesure: cat.Mesure, schemas: dict[str, Schema]) -> tuple[Schema, str]:
    return schemas[FEUILLE_DE_BASE[mesure.base]], COLONNE_DATE[mesure.base]


# ------------------------------------------------------------- Paramètres
def parametres(
    classeur: Workbook,
    situation: date,
    empreinte: EmpreinteJeu,
    seuil_exception: int,
    seuil_relance: int,
    fenetre: Fenetre,
) -> Worksheet:
    ws = classeur.create_sheet("Paramètres")
    theme.preparer(ws, largeur_b=34, colonnes=8, largeur=18)
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 92
    theme.bandeau(
        ws,
        "Paramètres, empreinte et méthode",
        "Les cellules bleues sur fond jaune sont les seules à modifier.",
        "Feuille de contrôle",
        colonnes=9,
    )

    theme.titre_section(ws, 7, 2, "Paramètres de calcul", 3)
    lignes: list[tuple[str, theme.ValeurCellule, str, str]] = [
        ("Date de situation", situation, theme.DATE,
         "Référence de tous les calculs d'âge et de retard. La modifier recalcule le rapport."),
        ("Début de la fenêtre", "=$C$8-364", theme.DATE, "Douze mois glissants."),
        ("Fin de la fenêtre", "=$C$8", theme.DATE, ""),
        ("Début de la comparaison", "=$C$8-729", theme.DATE, "Les douze mois précédents."),
        ("Fin de la comparaison", "=$C$8-365", theme.DATE, ""),
        ("Seuil de créance à risque", seuil_exception, theme.JOURS,
         "Ancienneté à partir de laquelle une créance passe en recouvrement."),
        ("Délai de relance visé", seuil_relance, theme.JOURS,
         "Au-delà, le taux de transformation chute nettement."),
    ]
    for decalage, (libelle, valeur, format_nombre, note) in enumerate(lignes):
        row = f.LIGNE_SITUATION + decalage
        cellule = ws.cell(row=row, column=2, value=libelle)
        cellule.font = Font(name=theme.POLICE, size=10, bold=True)
        cellule.border = theme.BORDURE
        if isinstance(valeur, str) and valeur.startswith("="):
            calculee = ws.cell(row=row, column=3, value=valeur)
            calculee.number_format = format_nombre
            calculee.border = theme.BORDURE
            calculee.font = Font(name=theme.POLICE, size=10)
            calculee.alignment = Alignment(horizontal="center")
        else:
            theme.saisie(ws, row, 3, valeur, format_nombre)
        commentaire = ws.cell(row=row, column=4, value=note)
        commentaire.font = Font(name=theme.POLICE, size=9, color=theme.GRIS)
        ws.row_dimensions[row].height = 18

    theme.titre_section(ws, 16, 2, "Conventions", 3)
    conventions = (
        ("Bleu sur fond jaune", "Cellule de saisie — modifiable", "0000FF", theme.JAUNE),
        ("Noir", "Résultat de formule — ne pas écraser", theme.ANTHRACITE, None),
        ("En-tête vert", "Colonne calculée ajoutée aux sources", theme.BLANC, theme.VERT),
        ("En-tête bleu nuit", "Donnée importée des CSV", theme.BLANC, theme.BLEU_NUIT),
    )
    for decalage, (libelle, note, couleur, fond) in enumerate(conventions):
        row = 17 + decalage
        cellule = ws.cell(row=row, column=2, value=libelle)
        cellule.font = Font(name=theme.POLICE, size=10, bold=True, color=couleur)
        if fond:
            cellule.fill = PatternFill("solid", fgColor=fond)
        cellule.border = theme.BORDURE
        commentaire = ws.cell(row=row, column=4, value=note)
        commentaire.font = Font(name=theme.POLICE, size=9, color=theme.GRIS)

    theme.titre_section(ws, 23, 2, "Empreinte du jeu de données", 3)
    theme.entetes(ws, 24, 2, ["Fichier", "Enregistrements", "SHA-256"])
    for decalage, fichier in enumerate(empreinte.fichiers):
        theme.ligne(
            ws,
            25 + decalage,
            2,
            [fichier.nom, fichier.enregistrements, fichier.sha256[:32] + "…"],
            [None, theme.NB, None],
            alterne=decalage % 2 == 1,
        )
    row = 25 + len(empreinte.fichiers)
    theme.ligne(
        ws,
        row,
        2,
        ["Ensemble", empreinte.enregistrements, grouper(empreinte.globale[:32])],
        [None, theme.NB, None],
        gras=True,
        fond=theme.BLEU_CLAIR,
    )

    debut_notes = row + 3
    theme.titre_section(ws, debut_notes - 1, 2, "Méthode", 3)
    notes = (
        "JEU DE DONNÉES FICTIF, généré pour démonstration. Aucune donnée réelle.",
        f"Période couverte : quatre exercices. Fenêtre d'analyse : {fenetre.libelle}.",
        "Le chiffre d'affaires est hors taxes ; l'encours en TTC, montant réellement dû.",
        "Le coût de revient d'une facture est celui de l'intervention liée : heures réelles "
        "fois coût horaire du technicien, plus matériel. Aucun frais de structure.",
        "La marge est donc une marge sur coûts directs, et non un résultat d'exploitation.",
        "Les taux somment numérateurs et dénominateurs AVANT de diviser. La moyenne "
        "de trois taux de marge n'est pas le taux de marge de trois agences.",
        "L'encours et le DSO sont des états à la date de situation : ils n'ont pas de version "
        "« période précédente », et aucune variation n'est affichée pour eux.",
        "L'empreinte établit que le rapport porte sur ces fichiers exacts. Elle "
        "n'établit aucune antériorité : il y faudrait un horodatage par un tiers.",
        "Toutes les cellules chiffrées sont des formules. Corriger une donnée source met à jour "
        "l'ensemble du rapport.",
    )
    for decalage, texte in enumerate(notes):
        cellule = ws.cell(row=debut_notes + decalage, column=2, value="• " + texte)
        cellule.font = Font(
            name=theme.POLICE, size=9, color=theme.ROUGE if decalage == 0 else theme.ANTHRACITE
        )
        ligne_note = debut_notes + decalage
        ws.merge_cells(
            start_row=ligne_note, start_column=2, end_row=ligne_note, end_column=4
        )
        ws.row_dimensions[debut_notes + decalage].height = 15
    return ws


# ---------------------------------------------------------- Détail mensuel
def detail_mensuel(
    classeur: Workbook, schemas: dict[str, Schema], mois: tuple[str, ...]
) -> Worksheet:
    ws = classeur.create_sheet("Détail mensuel")
    theme.preparer(ws, largeur_b=12, colonnes=11, largeur=15)
    theme.bandeau(
        ws,
        "Détail mensuel",
        "Source des graphiques de la synthèse. Une ligne par mois, sur toute la période.",
        "Données de série",
        colonnes=11,
    )

    devis, factures = schemas["Devis"], schemas["Factures"]
    theme.titre_section(ws, 7, 2, "Série mensuelle", 10)
    theme.entetes(
        ws,
        8,
        2,
        [
            "Mois",
            "CA facturé HT",
            "Coût de revient",
            "Marge brute",
            "Taux de marge",
            "Devis émis HT",
            "Devis arbitrés HT",
            "Devis gagnés HT",
            "Taux de transformation",
            "Encaissements TTC",
        ],
    )

    def somme(schema: Schema, champ: str, champ_mois: str, row: int) -> str:
        return f"=SUMIFS({schema.plage(champ)},{schema.plage(champ_mois)},$B{row})"

    for decalage, cle_mois in enumerate(mois):
        row = LIGNE_MOIS + decalage
        theme.ligne(
            ws,
            row,
            2,
            [
                cle_mois,
                somme(factures, "montant_ht", "mois", row),
                somme(factures, "cout_revient", "mois", row),
                f"=$C{row}-$D{row}",
                f'=IFERROR($E{row}/$C{row},"")',
                somme(devis, "montant_ht", "mois", row),
                somme(devis, "montant_arbitre", "mois", row),
                somme(devis, "ca_gagne", "mois", row),
                f'=IFERROR($I{row}/$H{row},"")',
                somme(factures, "montant_ttc", "mois_encaissement", row),
            ],
            [
                None,
                theme.EUR,
                theme.EUR,
                theme.EUR,
                theme.PCT,
                theme.EUR,
                theme.EUR,
                theme.EUR,
                theme.PCT,
                theme.EUR,
            ],
            alterne=decalage % 2 == 1,
        )

    total = LIGNE_MOIS + len(mois)
    derniere = total - 1
    theme.ligne(
        ws,
        total,
        2,
        ["Total"]
        + [f"=SUM({c}{LIGNE_MOIS}:{c}{derniere})" for c in "CD"]
        + [f"=$C{total}-$D{total}", f'=IFERROR($E{total}/$C{total},"")']
        + [f"=SUM({c}{LIGNE_MOIS}:{c}{derniere})" for c in "GHI"]
        + [f'=IFERROR($I{total}/$H{total},"")', f"=SUM($K{LIGNE_MOIS}:$K{derniere})"],
        [None, theme.EUR, theme.EUR, theme.EUR, theme.PCT, theme.EUR, theme.EUR,
         theme.EUR, theme.PCT, theme.EUR],
        gras=True,
        fond=theme.BLEU_CLAIR,
    )
    ws.print_area = f"A1:L{total}"
    return ws


# ------------------------------------------------------------------ Synthèse
def _carte_mesure(
    ws: Worksheet,
    row: int,
    col: int,
    mesure: cat.Mesure,
    schemas: dict[str, Schema],
    accent: str,
) -> None:
    schema, colonne_date = _schema_de(mesure, schemas)
    valeur = f.formule_mesure(mesure, schema, colonne_date, schemas=schemas)
    lettre = chr(ord("A") + col - 1)
    if mesure.comparable_entre_periodes:
        passee = f.formule_mesure(
            mesure, schema, colonne_date, f.DEBUT_COMPARAISON, f.FIN_COMPARAISON
        , schemas=schemas)
        sous_texte = f'=IFERROR(({valeur[1:]})/({passee[1:]})-1,"")'
        format_sous = theme.VARIATION
    else:
        sous_texte = "état à la date de situation"
        format_sous = None
    theme.carte(
        ws,
        row,
        col,
        mesure.libelle,
        valeur,
        f.format_unite(mesure),
        sous_texte,
        format_sous,
        accent,
    )
    _ = lettre


def synthese(
    classeur: Workbook,
    schemas: dict[str, Schema],
    selection: Selection,
    fenetre: Fenetre,
    metiers: tuple[str, ...],
    nb_mois: int,
) -> Worksheet:
    ws = classeur.create_sheet("Synthèse")
    theme.preparer(ws, largeur_b=17, colonnes=12, largeur=14)
    theme.bandeau(
        ws,
        "Rapport de pilotage — synthèse",
        f"Niveau {selection.niveau.value} · {fenetre.libelle} "
        f"({fenetre.debut:%d/%m/%Y} → {fenetre.fin:%d/%m/%Y})",
        "Bâti-Sud",
        colonnes=12,
    )

    accents = (theme.BLEU, theme.VERT, theme.ORANGE, theme.ROUGE)
    for indice, cle in enumerate(selection.socle):
        _carte_mesure(
            ws, 7, 2 + 3 * (indice % 4), cat.mesure(cle), schemas, accents[indice % 4]
        )
    for indice, indicateur in enumerate(selection.variables):
        _carte_mesure(
            ws,
            11,
            2 + 3 * (indice % 4),
            cat.mesure(indicateur.mesure),
            schemas,
            accents[indice % 4],
        )
    theme.note(
        ws,
        14,
        2,
        "Ligne du haut : socle imposé, comparable d'une édition à l'autre. "
        "Ligne du bas : indicateurs retenus pour cette période.",
    )

    theme.titre_section(ws, 16, 2, "Chiffre d'affaires facturé et taux de marge", 11)
    graphique = BarChart()
    graphique.type = "col"
    fin_mois = LIGNE_MOIS + nb_mois - 1
    detail = classeur["Détail mensuel"]
    graphique.add_data(
        Reference(detail, min_col=3, min_row=8, max_row=fin_mois), titles_from_data=True
    )
    graphique.set_categories(Reference(detail, min_col=2, min_row=LIGNE_MOIS, max_row=fin_mois))
    graphique.y_axis.numFmt = "#,##0"
    graphique.series[0].graphicalProperties.solidFill = theme.BLEU
    graphique.gapWidth = 40

    courbe = LineChart()
    courbe.add_data(
        Reference(detail, min_col=6, min_row=8, max_row=fin_mois), titles_from_data=True
    )
    courbe.y_axis.axId = 200
    courbe.y_axis.numFmt = "0%"
    courbe.series[0].graphicalProperties.line.solidFill = theme.ORANGE
    courbe.series[0].smooth = False
    graphique.y_axis.crosses = "autoZero"
    graphique += courbe
    graphique.height = 8.5
    graphique.width = 26
    graphique.title = None
    # `legend` est nullable sur un graphique openpyxl — ailleurs dans ce module,
    # on l'annule volontairement. La garde évite un accès à None si la valeur
    # par défaut venait à changer, et rend la nullabilité visible.
    if graphique.legend is not None:
        graphique.legend.position = "b"
    ws.add_chart(graphique, "B17")

    ligne_pont = _decomposition(ws, schemas, metiers, 35)
    _alertes(ws, selection, ligne_pont + 2)
    ws.print_area = f"A1:M{ligne_pont + 10}"
    return ws


def _somme_devis(schema: Schema, champ: str, debut: str, fin: str) -> str:
    """Corps de SUMIFS, sans le signe égal, pour composer des expressions."""
    date_ = schema.plage("date_emission")
    return f'SUMIFS({schema.plage(champ)},{date_},">="&{debut},{date_},"<="&{fin})'


def _decomposition(
    ws: Worksheet, schemas: dict[str, Schema], metiers: tuple[str, ...], depart: int
) -> int:
    """Ponts d'écart : volume, mix et prix côté facturation ; volume et
    transformation côté commercial.

    Les deux décompositions sont exactes — la somme des effets égale l'écart
    constaté — ce qui les rend vérifiables d'un coup d'œil par le lecteur.
    """
    factures, devis = schemas["Factures"], schemas["Devis"]
    theme.titre_section(ws, depart, 2, "Décomposition de l'écart de chiffre d'affaires", 6)
    theme.entetes(
        ws,
        depart + 1,
        2,
        ["Métier", "Affaires N-1", "Panier N-1", "Affaires N", "Panier N"],
    )

    def compte(debut: str, fin: str, row: int) -> str:
        return (
            f"=COUNTIFS({factures.plage('date_facture')},\">=\"&{debut},"
            f"{factures.plage('date_facture')},\"<=\"&{fin},"
            f"{factures.plage('metier')},$B{row})"
        )

    def montant(debut: str, fin: str, row: int) -> str:
        return (
            f"=SUMIFS({factures.plage('montant_ht')},{factures.plage('date_facture')},"
            f"\">=\"&{debut},{factures.plage('date_facture')},\"<=\"&{fin},"
            f"{factures.plage('metier')},$B{row})"
        )

    premiere = depart + 2
    for decalage, metier in enumerate(metiers):
        row = premiere + decalage
        theme.ligne(
            ws,
            row,
            2,
            [
                metier,
                compte(f.DEBUT_COMPARAISON, f.FIN_COMPARAISON, row),
                f'=IFERROR({montant(f.DEBUT_COMPARAISON, f.FIN_COMPARAISON, row)[1:]}/$C{row},0)',
                compte(f.DEBUT, f.FIN, row),
                f'=IFERROR({montant(f.DEBUT, f.FIN, row)[1:]}/$E{row},0)',
            ],
            [None, theme.NB, theme.EUR, theme.NB, theme.EUR],
            alterne=decalage % 2 == 1,
        )
    derniere = premiere + len(metiers) - 1
    total = derniere + 1
    theme.ligne(
        ws,
        total,
        2,
        [
            "Ensemble",
            f"=SUM($C{premiere}:$C{derniere})",
            f"=IFERROR(SUMPRODUCT($C{premiere}:$C{derniere},"
            f"$D{premiere}:$D{derniere})/$C{total},0)",
            f"=SUM($E{premiere}:$E{derniere})",
            f"=IFERROR(SUMPRODUCT($E{premiere}:$E{derniere},"
            f"$F{premiere}:$F{derniere})/$E{total},0)",
        ],
        [None, theme.NB, theme.EUR, theme.NB, theme.EUR],
        gras=True,
        fond=theme.BLEU_CLAIR,
    )

    pont = total + 2
    theme.titre_section(ws, pont, 2, "Ponts d'écart", 6)
    theme.entetes(ws, pont + 1, 2, ["Effet", "Montant", "Lecture"])
    #: Panier qu'on aurait constaté avec la structure de N et les prix de N-1.
    structure = (
        f"SUMPRODUCT($E{premiere}:$E{derniere},$D{premiere}:$D{derniere})/$E{total}"
    )
    effets = (
        (
            "Volume",
            f"=($E{total}-$C{total})*$D{total}",
            "Davantage d'affaires, au panier moyen de l'an dernier.",
        ),
        (
            "Mix",
            f"=IFERROR($E{total}*({structure}-$D{total}),0)",
            "Déformation de la répartition entre métiers, à prix constants.",
        ),
        (
            "Prix",
            f"=IFERROR($E{total}*($F{total}-{structure}),0)",
            "Variation du panier à structure constante.",
        ),
    )
    for decalage, (libelle, formule, lecture) in enumerate(effets):
        theme.ligne(
            ws,
            pont + 2 + decalage,
            2,
            [libelle, formule, lecture],
            [None, theme.EUR, None],
            alterne=decalage % 2 == 1,
        )
    row_total = pont + 5
    theme.ligne(
        ws,
        row_total,
        2,
        [
            "Écart total",
            f"=SUM($C{pont + 2}:$C{pont + 4})",
            "Égal, par construction, à la variation du chiffre d'affaires facturé.",
        ],
        [None, theme.EUR, None],
        gras=True,
        fond=theme.BLEU_CLAIR,
    )

    commercial = row_total + 2
    theme.titre_section(ws, commercial, 2, "Pont commercial — devis gagnés", 6)
    theme.entetes(ws, commercial + 1, 2, ["Effet", "Montant", "Lecture"])
    v0 = _somme_devis(devis, "montant_arbitre", f.DEBUT_COMPARAISON, f.FIN_COMPARAISON)
    v1 = _somme_devis(devis, "montant_arbitre", f.DEBUT, f.FIN)
    g0 = _somme_devis(devis, "ca_gagne", f.DEBUT_COMPARAISON, f.FIN_COMPARAISON)
    g1 = _somme_devis(devis, "ca_gagne", f.DEBUT, f.FIN)
    lignes_commerciales = (
        (
            "Volume devisé",
            f"=IFERROR(({v1}-{v0})*({g0})/({v0}),0)",
            "Effet du volume de devis émis, à taux de transformation constant.",
        ),
        (
            "Transformation",
            f"=IFERROR(({v1})*(({g1})/({v1})-({g0})/({v0})),0)",
            "Effet du taux de transformation, à volume constant.",
        ),
    )
    for decalage, (libelle, formule, lecture) in enumerate(lignes_commerciales):
        theme.ligne(
            ws,
            commercial + 2 + decalage,
            2,
            [libelle, formule, lecture],
            [None, theme.EUR, None],
            alterne=decalage % 2 == 1,
        )
    fin_commercial = commercial + 4
    theme.ligne(
        ws,
        fin_commercial,
        2,
        [
            "Écart total",
            f"=SUM($C{commercial + 2}:$C{commercial + 3})",
            "Égal à la variation du chiffre d'affaires gagné sur devis.",
        ],
        [None, theme.EUR, None],
        gras=True,
        fond=theme.BLEU_CLAIR,
    )
    ws.column_dimensions["D"].width = 62
    return fin_commercial


def _alertes(ws: Worksheet, selection: Selection, depart: int) -> None:
    theme.titre_section(ws, depart, 2, "Points d'attention", 11)
    messages: list[str] = []
    if selection.message_direction:
        messages.append(selection.message_direction)
    for indicateur in selection.variables:
        if indicateur.decision_attendue:
            libelle = cat.mesure(indicateur.mesure).libelle
            messages.append(f"{libelle} — {indicateur.decision_attendue}")
    for hypothese in selection.ecartees:
        messages.append(f"Écartée — {hypothese.enonce} : {hypothese.motif}")
    if not messages:
        messages.append("Aucun point d'attention n'a été relevé pour cette période.")

    for decalage, texte in enumerate(messages[:8]):
        row = depart + 1 + decalage
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
        cellule = ws.cell(row=row, column=2, value=texte)
        cellule.font = Font(name=theme.POLICE, size=9, color=theme.ANTHRACITE)
        cellule.alignment = Alignment(vertical="center", indent=1)
        cellule.fill = PatternFill(
            "solid", fgColor=theme.GRIS_CLAIR if decalage % 2 == 0 else theme.BLANC
        )
        ws.row_dimensions[row].height = 17


# --------------------------------------------------------------- Indicateurs
def indicateurs(
    classeur: Workbook,
    schemas: dict[str, Schema],
    selection: Selection,
    modalites: dict[str, tuple[str, ...]],
) -> Worksheet:
    ws = classeur.create_sheet("Indicateurs")
    theme.preparer(ws, largeur_b=26, colonnes=10, largeur=15)
    theme.bandeau(
        ws,
        "Indicateurs retenus",
        "Un bloc par indicateur : ventilation, comparaison et décision attendue.",
        "Détail",
        colonnes=10,
    )

    row = 7
    for indicateur in selection.variables:
        row = _bloc_indicateur(ws, schemas, indicateur, modalites, row) + 3
    ws.print_area = f"A1:K{row}"
    return ws


def _bloc_indicateur(
    ws: Worksheet,
    schemas: dict[str, Schema],
    indicateur: Indicateur,
    modalites: dict[str, tuple[str, ...]],
    depart: int,
) -> int:
    mesure = cat.mesure(indicateur.mesure)
    schema, colonne_date = _schema_de(mesure, schemas)
    libelle = mesure.libelle
    if indicateur.dimension:
        libelle += f" par {cat.dimension(indicateur.dimension).libelle.lower()}"
    theme.titre_section(ws, depart, 2, libelle, 7)

    if indicateur.pourquoi:
        theme.note(ws, depart + 1, 2, indicateur.pourquoi, colonnes=7)
    entete = depart + 2

    if indicateur.dimension is None:
        theme.entetes(ws, entete, 2, ["Période", "Valeur", "Effectif"])
        for decalage, (etiquette, debut, fin) in enumerate(
            (("Période", f.DEBUT, f.FIN), ("Précédente", f.DEBUT_COMPARAISON, f.FIN_COMPARAISON))
        ):
            theme.ligne(
                ws,
                entete + 1 + decalage,
                2,
                [
                    etiquette,
                    f.formule_mesure(mesure, schema, colonne_date, debut, fin, schemas=schemas),
                    f.formule_effectif(schema, colonne_date, debut, fin),
                ],
                [None, f.format_unite(mesure), theme.NB],
                alterne=decalage % 2 == 1,
            )
        return entete + 2

    dimension = cat.dimension(indicateur.dimension)
    theme.entetes(
        ws,
        entete,
        2,
        ["Modalité", "Valeur", "Effectif", "Période précédente", "Écart", "Part du total"],
    )
    liste = modalites[indicateur.dimension]
    for decalage, modalite in enumerate(liste):
        row = entete + 1 + decalage
        critere = (dimension.colonne, f'"{modalite}"')
        actuelle = f.formule_mesure(
            mesure, schema, colonne_date, dimension=critere, schemas=schemas
        )
        passee = f.formule_mesure(
            mesure, schema, colonne_date, f.DEBUT_COMPARAISON, f.FIN_COMPARAISON, critere
        , schemas=schemas)
        ecart = (
            f.formule_variation(f"$C{row}", f"$E{row}")
            if mesure.comparable_entre_periodes
            else '="état"'
        )
        numerateur = mesure.agregat.numerateur if mesure.agregat else "montant_ht"
        part = (
            f"=IFERROR(SUMIFS({schema.plage(numerateur)},"
            f"{schema.plage(colonne_date)},\">=\"&{f.DEBUT},"
            f"{schema.plage(colonne_date)},\"<=\"&{f.FIN},"
            f"{schema.plage(dimension.colonne)},\"{modalite}\")"
            f"/SUMIFS({schema.plage(numerateur)},"
            f"{schema.plage(colonne_date)},\">=\"&{f.DEBUT},"
            f'{schema.plage(colonne_date)},"<="&{f.FIN}),"")'
        )
        theme.ligne(
            ws,
            row,
            2,
            [
                modalite,
                actuelle,
                f.formule_effectif(schema, colonne_date, dimension=critere),
                passee,
                ecart,
                part,
            ],
            [
                None,
                f.format_unite(mesure),
                theme.NB,
                f.format_unite(mesure),
                theme.VARIATION if mesure.comparable_entre_periodes else None,
                theme.PCT,
            ],
            alterne=decalage % 2 == 1,
        )

    fin = entete + len(liste)
    theme.ligne(
        ws,
        fin + 1,
        2,
        [
            "Ensemble",
            f.formule_mesure(mesure, schema, colonne_date, schemas=schemas),
            f.formule_effectif(schema, colonne_date),
            f.formule_mesure(
                mesure,
                schema,
                colonne_date,
                f.DEBUT_COMPARAISON,
                f.FIN_COMPARAISON,
                schemas=schemas,
            ),
            f.formule_variation(f"$C{fin + 1}", f"$E{fin + 1}")
            if mesure.comparable_entre_periodes
            else '="état"',
            "=1",
        ],
        [None, f.format_unite(mesure), theme.NB, f.format_unite(mesure),
         theme.VARIATION if mesure.comparable_entre_periodes else None, theme.PCT],
        gras=True,
        fond=theme.BLEU_CLAIR,
    )
    if indicateur.seuil_alerte is not None:
        theme.feux(
            ws,
            f"C{entete + 1}:C{fin}",
            indicateur.seuil_alerte,
            plus_haut_mieux=mesure.sens is cat.Sens.HAUT,
        )

    graphique = BarChart()
    graphique.type = "bar"
    graphique.add_data(Reference(ws, min_col=3, min_row=entete, max_row=fin), titles_from_data=True)
    graphique.set_categories(Reference(ws, min_col=2, min_row=entete + 1, max_row=fin))
    graphique.series[0].graphicalProperties.solidFill = theme.BLEU
    graphique.gapWidth = 55
    graphique.height = HAUTEUR_BARRE_CM * len(liste) + HAUTEUR_SOCLE_CM
    graphique.width = 13
    graphique.title = None
    graphique.legend = None
    ws.add_chart(graphique, f"I{entete}")

    ligne_decision = fin + 2
    if indicateur.decision_attendue:
        theme.note(ws, ligne_decision, 2, "Décision attendue : " + indicateur.decision_attendue, 7)
        derniere = ligne_decision
    else:
        derniere = fin + 1

    # Le bloc suivant doit démarrer sous le graphique, pas sous le tableau : un
    # graphique de cinq barres occupe une quinzaine de lignes, là où le tableau
    # n'en prend que sept, et les blocs se chevauchaient silencieusement.
    bas_graphique = entete + ceil(graphique.height / HAUTEUR_LIGNE_CM)
    return max(derniere, bas_graphique)


# ------------------------------------------------------------------ Criblage
def criblage(classeur: Workbook, resultat: Criblage, selection: Selection) -> Worksheet:
    """Ce qui a été examiné, retenu et écarté.

    Valeurs figées : cette feuille constate un processus passé, pas un état des
    données. La recalculer n'aurait pas de sens.
    """
    ws = classeur.create_sheet("Ce qui a été regardé")
    theme.preparer(ws, largeur_b=44, colonnes=8, largeur=16)
    theme.bandeau(
        ws,
        "Ce qui a été regardé",
        f"{resultat.explores} croisements examinés, {len(resultat.evalues)} recevables, "
        f"{len(resultat.rejetes)} écartés avant scoring.",
        "Traçabilité",
        colonnes=8,
    )

    theme.titre_section(ws, 7, 2, "Candidats les mieux placés", 7)
    theme.entetes(
        ws,
        8,
        2,
        [
            "Croisement",
            "Matérialité (€/an)",
            "Dispersion",
            "Stabilité",
            "Monotonie",
            "Score",
            "Retenu",
        ],
    )
    retenus = {i.cle for i in selection.variables}
    for decalage, candidat in enumerate(resultat.retenus):
        theme.ligne(
            ws,
            9 + decalage,
            2,
            [
                candidat.libelle,
                candidat.scores.materialite_euros,
                candidat.scores.dispersion,
                candidat.scores.stabilite,
                candidat.scores.monotonie,
                candidat.score_global,
                "oui" if candidat.cle in retenus else "",
            ],
            [None, theme.EUR, theme.PCT, theme.NB1, theme.NB1, theme.NB1, None],
            alterne=decalage % 2 == 1,
        )
    theme.barres(ws, f"C9:C{8 + len(resultat.retenus)}", theme.BLEU)

    depart = 10 + len(resultat.retenus)
    theme.titre_section(ws, depart, 2, "Écartés avant scoring", 7)
    theme.entetes(ws, depart + 1, 2, ["Croisement", "Motif"])
    for decalage, candidat in enumerate(resultat.rejetes):
        theme.ligne(
            ws,
            depart + 2 + decalage,
            2,
            [candidat.libelle, candidat.motif_rejet],
            [None, None],
            alterne=decalage % 2 == 1,
        )
    ws.column_dimensions["C"].width = 58

    if selection.ecartees:
        base = depart + 3 + len(resultat.rejetes)
        theme.titre_section(ws, base, 2, "Hypothèses explorées puis abandonnées", 7)
        theme.entetes(ws, base + 1, 2, ["Hypothèse", "Motif d'abandon"])
        for decalage, hypothese in enumerate(selection.ecartees):
            theme.ligne(
                ws,
                base + 2 + decalage,
                2,
                [hypothese.enonce, hypothese.motif],
                [None, None],
                alterne=decalage % 2 == 1,
            )
    return ws


# ------------------------------------------------------- Files de travail
def files_de_travail(
    classeur: Workbook,
    files: tuple[File, ...],
    situation: date,
    empreinte: EmpreinteJeu,
) -> Worksheet:
    """Le livrable du niveau opérationnel : des listes à traiter, pas des cartes.

    Valeurs figées et non des formules, contrairement au reste du classeur. Une
    file de travail constate un état à un instant : la recalculer trois semaines
    plus tard changerait les priorités sous les yeux de qui la traite, et ferait
    rappeler des clients déjà réglés. C'est le seul endroit du classeur où le
    figement est la bonne réponse.
    """
    ws = classeur.create_sheet("Files de travail")
    theme.preparer(ws, largeur_b=16, colonnes=9, largeur=15)
    theme.bandeau(
        ws,
        "Files de travail",
        f"Situation au {situation:%d/%m/%Y} · listes triées par priorité, "
        "à traiter dans l'ordre",
        "Opérationnel",
        colonnes=9,
    )

    ligne_courante = 7
    for file in files:
        theme.titre_section(ws, ligne_courante, 2, file.titre, 8)
        theme.note(
            ws,
            ligne_courante + 1,
            2,
            f"{file.total_candidats} à traiter, {len(file.taches)} affichées · "
            f"{file.seuil_libelle} · total affiché {file.montant_total:,.0f} "
            f"{file.unite_montant}".replace(",", " "),
            8,
        )
        entete = ligne_courante + 2
        theme.entetes(
            ws,
            entete,
            2,
            ["Référence", "Objet", f"Montant ({file.unite_montant})", "Ancienneté", "Motif"],
        )
        for decalage, tache in enumerate(file.taches):
            theme.ligne(
                ws,
                entete + 1 + decalage,
                2,
                [
                    tache.reference,
                    tache.libelle,
                    round(tache.montant, 2),
                    tache.anciennete_jours,
                    tache.motif,
                ],
                [None, None, theme.EUR, theme.NB, None],
                alterne=decalage % 2 == 1,
            )
        ligne_courante = entete + len(file.taches) + 3

    theme.note(
        ws,
        ligne_courante,
        2,
        "Ces listes sont figées à la date de situation. Les réexporter en JSON "
        "permet de les pousser dans la file d'un agent de relance : "
        f"empreinte du jeu {grouper(empreinte.globale)[:39]}…",
        8,
    )
    ws.print_area = f"A1:J{ligne_courante}"
    return ws
