"""Palette, formats de nombre et primitives de mise en page.

Un seul endroit décide de l'apparence. Les modules qui construisent les feuilles
ne manipulent jamais une couleur ni une police directement : ils appellent
`bandeau`, `carte`, `entetes`, `ligne`. Changer l'identité visuelle du rapport
tient alors dans ce fichier.

Les conventions de couleur suivent l'usage des modèles financiers : bleu sur
fond jaune pour une saisie, noir pour un résultat de formule. Un lecteur habitué
sait immédiatement ce qu'il peut modifier sans rien casser.
"""

from __future__ import annotations

from collections.abc import Sequence
from datetime import date

from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

ANTHRACITE = "1B2733"
BLEU = "1F6FB2"
BLEU_CLAIR = "E8F1F9"
BLEU_NUIT = "17394F"
VERT = "1E8A6E"
VERT_CLAIR = "E6F4EF"
ORANGE = "D9762B"
ROUGE = "C0392B"
ROUGE_CLAIR = "FBEAE8"
GRIS = "6B7A88"
GRIS_CLAIR = "F2F5F7"
GRIS_LIGNE = "D6DDE3"
BLANC = "FFFFFF"
JAUNE = "FFF3C4"

POLICE = "Arial"

EUR = '#,##0 " €";-#,##0 " €";"-"'
EUR2 = '#,##0.00 " €";-#,##0.00 " €";"-"'
PCT = "0.0%"
NB = '#,##0;-#,##0;"-"'
NB1 = '#,##0.0;-#,##0.0;"-"'
JOURS = '#,##0 " j";-#,##0 " j";"-"'
DATE = "DD/MM/YYYY"
VARIATION = '+0.0%;-0.0%;"stable"'
POINTS = '+0.0" pt";-0.0" pt";"-"'

#: Format de nombre associé à chaque unité du catalogue.
FORMAT_UNITE: dict[str, str] = {"€": EUR, "%": PCT, "j": JOURS, "h": NB1, "": NB}

_FIN = Side(style="thin", color=GRIS_LIGNE)
BORDURE = Border(left=_FIN, right=_FIN, top=_FIN, bottom=_FIN)
_SOUS_TITRE = Side(style="medium", color=ANTHRACITE)

LARGEUR_BANDEAU = 74


#: Ce qu'une cellule accepte réellement. `object` était trop large : openpyxl
#: refuse d'y écrire n'importe quoi, et le disait sans qu'on l'entende.
ValeurCellule = str | float | int | date | None


def preparer(ws: Worksheet, largeur_b: int = 20, colonnes: int = 12, largeur: int = 14) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2.2
    ws.column_dimensions["B"].width = largeur_b
    for indice in range(3, 2 + colonnes):
        ws.column_dimensions[get_column_letter(indice)].width = largeur
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True



def bandeau(ws: Worksheet, titre: str, sous_titre: str, mention: str, colonnes: int = 13) -> None:
    for ligne_ in range(1, 6):
        for colonne in range(1, colonnes + 2):
            ws.cell(row=ligne_, column=colonne).fill = PatternFill("solid", fgColor=ANTHRACITE)
    for ligne_, hauteur in ((1, 8), (2, 16), (3, 28), (4, 16), (5, 10)):
        ws.row_dimensions[ligne_].height = hauteur

    marque = "B Â T I - S U D    ·    G R O U P E   A R T I S A N A L"
    entete = ws.cell(row=2, column=2, value=marque)
    entete.font = Font(name=POLICE, size=8, bold=True, color="7FB3D5")
    principal = ws.cell(row=3, column=2, value=titre)
    principal.font = Font(name=POLICE, size=17, bold=True, color=BLANC)
    secondaire = ws.cell(row=4, column=2, value=sous_titre)
    secondaire.font = Font(name=POLICE, size=9, color="A9BAC6")
    droite = ws.cell(row=3, column=colonnes - 1, value=mention)
    droite.font = Font(name=POLICE, size=9, color="7FB3D5")
    droite.alignment = Alignment(horizontal="right")


def titre_section(ws: Worksheet, row: int, col: int, texte: str, span: int = 6) -> None:
    cellule = ws.cell(row=row, column=col, value=texte)
    cellule.font = Font(name=POLICE, size=11, bold=True, color=ANTHRACITE)
    for indice in range(col, col + span):
        ws.cell(row=row, column=indice).border = Border(bottom=_SOUS_TITRE)
    ws.row_dimensions[row].height = 22


def carte(
    ws: Worksheet,
    row: int,
    col: int,
    label: str,
    formule: str,
    format_nombre: str,
    sous_texte: str | None = None,
    format_sous_texte: str | None = None,
    accent: str = BLEU,
) -> None:
    """Carte d'indicateur : trois colonnes de large, trois lignes de haut."""
    fin = col + 2
    for decalage in range(3):
        ws.merge_cells(
            start_row=row + decalage, start_column=col, end_row=row + decalage, end_column=fin
        )
    for ligne_ in range(row, row + 3):
        for colonne in range(col, fin + 1):
            cellule = ws.cell(row=ligne_, column=colonne)
            cellule.fill = PatternFill("solid", fgColor=GRIS_CLAIR)
            cellule.border = Border(
                top=Side(style="thick", color=accent) if ligne_ == row else None,
                left=_FIN if colonne == col else None,
                right=_FIN if colonne == fin else None,
                bottom=_FIN if ligne_ == row + 2 else None,
            )

    titre = ws.cell(row=row, column=col, value=label.upper())
    titre.font = Font(name=POLICE, size=8, bold=True, color=GRIS)
    titre.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    valeur = ws.cell(row=row + 1, column=col, value=formule)
    valeur.font = Font(name=POLICE, size=18, bold=True, color=ANTHRACITE)
    valeur.number_format = format_nombre
    valeur.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    complement = ws.cell(row=row + 2, column=col, value=sous_texte)
    complement.font = Font(name=POLICE, size=8, color=GRIS)
    if format_sous_texte:
        complement.number_format = format_sous_texte
    complement.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    for ligne_, hauteur in ((row, 18), (row + 1, 30), (row + 2, 16)):
        ws.row_dimensions[ligne_].height = hauteur


def entetes(ws: Worksheet, row: int, col: int, libelles: list[str], fond: str = BLEU_NUIT) -> None:
    for decalage, libelle in enumerate(libelles):
        cellule = ws.cell(row=row, column=col + decalage, value=libelle)
        cellule.font = Font(name=POLICE, size=9, bold=True, color=BLANC)
        cellule.fill = PatternFill("solid", fgColor=fond)
        cellule.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cellule.border = BORDURE
    ws.row_dimensions[row].height = 30


def ligne(
    ws: Worksheet,
    row: int,
    col: int,
    valeurs: Sequence[ValeurCellule],
    formats: list[str | None],
    gras: bool = False,
    fond: str | None = None,
    alterne: bool = False,
) -> None:
    for decalage, (valeur, format_nombre) in enumerate(zip(valeurs, formats, strict=True)):
        cellule = ws.cell(row=row, column=col + decalage, value=valeur)
        cellule.font = Font(name=POLICE, size=9, bold=gras, color=ANTHRACITE)
        if fond:
            cellule.fill = PatternFill("solid", fgColor=fond)
        elif alterne:
            cellule.fill = PatternFill("solid", fgColor=GRIS_CLAIR)
        cellule.border = BORDURE
        if format_nombre:
            cellule.number_format = format_nombre
        cellule.alignment = Alignment(
            horizontal="left" if decalage == 0 else "right", vertical="center"
        )
    ws.row_dimensions[row].height = 17


def barres(ws: Worksheet, plage: str, couleur: str = BLEU) -> None:
    ws.conditional_formatting.add(
        plage,
        DataBarRule(
            start_type="num", start_value=0, end_type="max", color=couleur, showValue=True
        ),
    )


def feux(ws: Worksheet, plage: str, seuil: float, plus_haut_mieux: bool = True) -> None:
    """Vert du bon côté du seuil, rouge de l'autre."""
    bon, mauvais = (
        ("greaterThanOrEqual", "lessThan")
        if plus_haut_mieux
        else ("lessThanOrEqual", "greaterThan")
    )
    ws.conditional_formatting.add(
        plage,
        CellIsRule(
            operator=bon,
            formula=[str(seuil)],
            font=Font(name=POLICE, size=9, bold=True, color="14654F"),
            fill=PatternFill("solid", fgColor=VERT_CLAIR),
        ),
    )
    ws.conditional_formatting.add(
        plage,
        CellIsRule(
            operator=mauvais,
            formula=[str(seuil)],
            font=Font(name=POLICE, size=9, bold=True, color="93261B"),
            fill=PatternFill("solid", fgColor=ROUGE_CLAIR),
        ),
    )


def saisie(
    ws: Worksheet, row: int, col: int, valeur: ValeurCellule, format_nombre: str
) -> None:
    """Cellule modifiable : bleu sur fond jaune, comme dans un modèle financier."""
    cellule = ws.cell(row=row, column=col, value=valeur)
    cellule.font = Font(name=POLICE, size=10, bold=True, color="0000FF")
    cellule.fill = PatternFill("solid", fgColor=JAUNE)
    cellule.number_format = format_nombre
    cellule.border = BORDURE
    cellule.alignment = Alignment(horizontal="center")


def note(ws: Worksheet, row: int, col: int, texte: str, colonnes: int = 11) -> None:
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + colonnes)
    cellule = ws.cell(row=row, column=col, value=texte)
    cellule.font = Font(name=POLICE, size=9, color=GRIS)
    cellule.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 15
