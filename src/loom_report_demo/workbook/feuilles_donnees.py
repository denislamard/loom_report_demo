"""Écriture des sept feuilles de données et de leurs colonnes calculées.

Les données sources sont écrites telles quelles ; les colonnes calculées sont des
formules. Le classeur reste donc vivant : corriger un montant dans la feuille
Devis met à jour la marge, le taux de transformation, le graphique mensuel et les
cartes de la synthèse, sans rien relancer.

Les en-têtes distinguent les deux : bleu nuit pour une donnée importée, vert pour
une colonne calculée. Un lecteur sait tout de suite ce qu'il peut corriger.
"""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from loom_report_demo import paths
from loom_report_demo.workbook import theme
from loom_report_demo.workbook.formules import COLONNES_CALCULEES, substitutions_globales
from loom_report_demo.workbook.schema import FEUILLES, Schema

#: Champs à convertir, par suffixe ou par nom. Le reste reste du texte.
_DATES = ("date_",)
_ENTIERS = {"quantite", "rang", "nb_relances", "delai_contractuel_j", "delai_1ere_relance_j"}
_REELS = {
    "montant_ht",
    "montant_ttc",
    "taux_tva",
    "duree_min",
    "heures_devisees",
    "heures_reelles",
    "cout_horaire",
    "cout_main_oeuvre",
    "cout_materiel",
    "prix_unitaire_ht",
    "marge_cible_pct",
    "heures_standard",
    "taux_facturation_horaire",
}
_FORMATS = {
    "montant_ht": theme.EUR,
    "montant_ttc": theme.EUR,
    "cout_main_oeuvre": theme.EUR,
    "cout_materiel": theme.EUR,
    "prix_unitaire_ht": theme.EUR,
    "cout_horaire": theme.EUR2,
    "taux_facturation_horaire": theme.EUR2,
    "taux_tva": theme.PCT,
    "marge_cible_pct": theme.PCT,
    "heures_devisees": theme.NB1,
    "heures_reelles": theme.NB1,
    "heures_standard": theme.NB1,
    "duree_min": theme.NB1,
}
_LARGEURS = {
    "libelle_prestation": 32,
    "libelle": 38,
    "nom_client": 28,
    "nom_technicien": 20,
    "ville": 16,
    "statut_intervention": 15,
    "canal_acquisition": 18,
    "profil_paiement": 16,
}


def _convertir(valeur: str, champ: str) -> object:
    if valeur == "":
        return None
    if champ.startswith(_DATES):
        return datetime.strptime(valeur, "%Y-%m-%d").date()
    if champ in _ENTIERS:
        return int(valeur)
    if champ in _REELS:
        return float(valeur)
    return valeur


def _entete(ws: Worksheet, colonne: int, texte: str, fond: str) -> None:
    cellule = ws.cell(row=1, column=colonne, value=texte)
    cellule.font = Font(name=theme.POLICE, size=9, bold=True, color=theme.BLANC)
    cellule.fill = PatternFill("solid", fgColor=fond)
    cellule.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cellule.border = theme.BORDURE


def _libelle(champ: str) -> str:
    return champ.replace("_", " ").capitalize()


def ecrire_feuilles(
    classeur: Workbook, schemas: dict[str, Schema], dossier: Path | None = None
) -> None:
    """Écrit les sept feuilles, données puis colonnes calculées."""
    source = dossier or paths.donnees()
    substitutions = substitutions_globales(schemas)

    for feuille, fichier in FEUILLES.items():
        with (source / fichier).open(encoding="utf-8") as flux:
            lecteur = csv.reader(flux, delimiter=";")
            champs = next(lecteur)
            lignes = list(lecteur)

        ws = classeur.create_sheet(feuille)
        ws.sheet_view.showGridLines = False

        for indice, champ in enumerate(champs, start=1):
            _entete(ws, indice, _libelle(champ), theme.BLEU_NUIT)
            ws.column_dimensions[get_column_letter(indice)].width = _LARGEURS.get(
                champ, max(10, min(20, len(champ) + 4))
            )
        ws.row_dimensions[1].height = 30

        for numero, donnees in enumerate(lignes, start=2):
            for indice, (valeur, champ) in enumerate(zip(donnees, champs, strict=True), start=1):
                cellule = ws.cell(row=numero, column=indice, value=_convertir(valeur, champ))
                cellule.font = Font(name=theme.POLICE, size=9)
                cellule.border = theme.BORDURE
                if champ in _FORMATS:
                    cellule.number_format = _FORMATS[champ]
                elif champ.startswith(_DATES):
                    cellule.number_format = theme.DATE
                if champ in _ENTIERS or champ in _REELS:
                    cellule.alignment = Alignment(horizontal="right")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(champs))}{len(lignes) + 1}"

        schema = schemas[feuille]
        lettres = {champ: schema.ancre(champ) for champ in schema.champs}
        depart = len(champs) + 1
        for decalage, colonne in enumerate(COLONNES_CALCULEES[feuille]):
            indice = depart + decalage
            _entete(ws, indice, colonne.entete, theme.VERT)
            ws.column_dimensions[get_column_letter(indice)].width = colonne.largeur
            for numero in range(2, len(lignes) + 2):
                cellule = ws.cell(
                    row=numero,
                    column=indice,
                    value=colonne.gabarit.format(r=numero, **lettres, **substitutions),
                )
                cellule.font = Font(name=theme.POLICE, size=9)
                cellule.border = theme.BORDURE
                cellule.alignment = Alignment(horizontal="right")
                if colonne.format:
                    cellule.number_format = colonne.format
