"""Construction du classeur Excel.

`openpyxl` uniquement : ce paquet ne connaît ni `loom_ia`, ni le réseau. Il
reçoit une sélection d'indicateurs et produit un fichier — que la sélection
vienne d'une fixture écrite à la main ou du modèle ne change rien pour lui.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from loom_report_demo import paths
from loom_report_demo.analysis import catalogue as cat
from loom_report_demo.analysis.cadrages import fenetre
from loom_report_demo.analysis.chargement import (
    SEUIL_EXCEPTION_JOURS,
    SEUIL_RELANCE_RAPIDE_JOURS,
    Donnees,
)
from loom_report_demo.analysis.chargement import (
    donnees as charger_donnees,
)
from loom_report_demo.analysis.criblage import cribler
from loom_report_demo.analysis.faisabilite import produit_des_files
from loom_report_demo.analysis.files import construire_files
from loom_report_demo.fingerprint import empreinte_jeu
from loom_report_demo.niveaux import Niveau
from loom_report_demo.workbook import tableaux
from loom_report_demo.workbook.feuilles_donnees import ecrire_feuilles
from loom_report_demo.workbook.formules import (
    COLONNES_CALCULEES,
    COLONNES_STRATEGIQUES,
)
from loom_report_demo.workbook.schema import ColonneCalculee, construire_schemas
from loom_report_demo.workbook.selection import Selection

#: Ordre d'apparition des onglets : restitution d'abord, sources ensuite.
ORDRE = (
    "Synthèse",
    "Indicateurs",
    "Détail mensuel",
    "Ce qui a été regardé",
    "Paramètres",
    "Devis",
    "Relances",
    "Factures",
    "Interventions",
    "Clients",
    "Techniciens",
    "Catalogue",
)

COULEURS_ONGLET = {
    "Synthèse": "1B2733",
    "Indicateurs": "1F6FB2",
    "Détail mensuel": "1E8A6E",
    "Ce qui a été regardé": "D9762B",
    "Paramètres": "6B7A88",
}


def _modalites(source: Donnees, selection: Selection) -> dict[str, tuple[str, ...]]:
    """Libellés des modalités, dans l'ordre du catalogue quand il est imposé.

    Seuls les libellés sont extraits des données : les valeurs, elles, seront
    calculées par le classeur.
    """
    resultat: dict[str, tuple[str, ...]] = {}
    for indicateur in selection.variables:
        if indicateur.dimension is None:
            continue
        dimension = cat.dimension(indicateur.dimension)
        mesure = cat.mesure(indicateur.mesure)
        colonne = source.table(mesure.base)[dimension.colonne]
        presentes = list(dict.fromkeys(str(x) for x in colonne.dropna()))
        if dimension.ordonnee and dimension.modalites:
            ordonnees = [m for m in dimension.modalites if m in presentes]
            presentes = ordonnees + [m for m in presentes if m not in dimension.modalites]
        else:
            presentes.sort()
        resultat[indicateur.dimension] = tuple(presentes)
    return resultat


def _colonnes(niveau: Niveau) -> dict[str, tuple[ColonneCalculee, ...]]:
    """Colonnes calculées à écrire, selon le niveau.

    Les colonnes d'appui du stratégique coûtent cher — chaque ligne balaie celles
    qui la précèdent — et ne servent qu'à deux mesures. On ne les écrit donc que
    là où elles sont utiles.
    """
    if niveau is not Niveau.STRATEGIQUE:
        return COLONNES_CALCULEES
    fusion = {feuille: tuple(colonnes) for feuille, colonnes in COLONNES_CALCULEES.items()}
    for feuille, ajouts in COLONNES_STRATEGIQUES.items():
        fusion[feuille] = fusion.get(feuille, ()) + ajouts
    return fusion


def construire(
    selection: Selection,
    destination: Path,
    source: Donnees | None = None,
    dossier_donnees: Path | None = None,
    strict: bool = True,
) -> Path:
    """Produit le classeur et rend son chemin.

    `strict` vaut faux quand l'humain a retiré un indicateur après arbitrage : la
    sélection compte alors moins de variables que le niveau n'en prévoit, ce qui
    est légitime — jamais plus.
    """
    selection.valider(strict=strict)
    jeu = source if source is not None else charger_donnees(dossier_donnees)
    schemas = construire_schemas(_colonnes(selection.niveau), dossier_donnees)
    borne = fenetre(selection.cadrage, jeu.situation, jeu.debut)
    mois = tuple(
        sorted(
            {str(cle) for cle in jeu.factures["mois"]}
            | {str(cle) for cle in jeu.devis["mois"]}
        )
    )
    metiers = tuple(sorted(str(x) for x in jeu.factures["metier"].dropna().unique()))
    empreinte = empreinte_jeu([paths.csv_source(n) for n in paths.FICHIERS_DONNEES])

    classeur = Workbook()
    # `Workbook()` crée une feuille vide : on la retire par son index plutôt
    # que par `active`, qui peut être nulle et n'est pas typée comme feuille.
    classeur.remove(classeur.worksheets[0])

    ecrire_feuilles(classeur, schemas, dossier_donnees)
    tableaux.parametres(
        classeur, jeu.situation, empreinte, SEUIL_EXCEPTION_JOURS, SEUIL_RELANCE_RAPIDE_JOURS, borne
    )
    tableaux.detail_mensuel(classeur, schemas, mois)
    tableaux.synthese(classeur, schemas, selection, borne, metiers, len(mois))
    tableaux.indicateurs(classeur, schemas, selection, _modalites(jeu, selection))
    tableaux.criblage(classeur, cribler(jeu, selection.niveau), selection)

    ordre: list[str] = list(ORDRE)
    if produit_des_files(selection.niveau):
        tableaux.files_de_travail(
            classeur, construire_files(jeu, selection.seuils()), jeu.situation, empreinte
        )
        ordre.insert(1, "Files de travail")

    # Réordonnancement par l'API publique : écraser `_sheets` fonctionnait, mais
    # reposait sur un détail d'implémentation d'openpyxl.
    for position, nom in enumerate(ordre):
        actuelle = classeur.sheetnames.index(nom)
        if actuelle != position:
            classeur.move_sheet(nom, offset=position - actuelle)
    for nom in ordre:
        classeur[nom].sheet_properties.tabColor = COULEURS_ONGLET.get(nom, "B8C4CE")
    classeur["Synthèse"].sheet_view.tabSelected = True
    classeur.active = 0

    destination.parent.mkdir(parents=True, exist_ok=True)
    classeur.save(destination)
    return destination
